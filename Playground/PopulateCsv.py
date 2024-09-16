# import os
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image
#
# badgeID = []
# fileNames = []
# imagePath = "/Users/u75530/Documents/GitHub/Python_Django/Playground/Photopath"
# xlsxFilePath = "/Users/u75530/Documents/GitHub/Python_Django/Playground/ID.xlsx"
#
# # Read the filenames from the Photopath folder
# try:
#     readFileNames = os.listdir(imagePath)
#
#     # Assuming file names are badge IDs with an extension like "12345.jpg"
#     for file in readFileNames:
#         fileNames.append(file[0:5])  # Extract badge ID from the first 5 characters
#
#     print(fileNames)
#     print(len(fileNames))
#
# except Exception as error:
#     print(f"Error reading file names: {error}")
#
# # Open the Excel workbook and read data
# try:
#     workbook = load_workbook(xlsxFilePath)
#     sheet = workbook.active  # Assuming data is in the first sheet
#
#     # Iterate through the rows of the Excel file starting from row 2 (if headers are in row 1)
#     for row in range(2, sheet.max_row + 1):
#         badge_id = str(sheet[f"A{row}"].value)  # Read BadgeID from column A
#
#         if badge_id in fileNames:
#             # Construct the full image path for the matched badge ID
#             image_file = f"{imagePath}/{badge_id}.jpg"
#
#             # Check if the image file exists
#             if os.path.exists(image_file):
#                 # Insert the image into column B of the current row
#                 img = Image(image_file)
#                 img.height = 100  # Adjust height if necessary
#                 img.width = 100   # Adjust width if necessary
#                 sheet.add_image(img, f"B{row}")
#
#     # Save the workbook after modifications
#     workbook.save(xlsxFilePath)
#     print("Excel file updated with images.")
#
# except Exception as error:
#     print(f"Error processing Excel file: {error}")

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

# Create a new Excel workbook and select the active worksheet
workbook = Workbook()
sheet = workbook.active

# Set the column width and row height
column_letter = 'B'
row_number = 2

# Set the size of the cell where the image will be placed
sheet.column_dimensions[column_letter].width = 20  # Set column width
sheet.row_dimensions[row_number].height = 100      # Set row height

# Load the image
img_path = 'python.jpg'
img = Image(img_path)

# Get the cell dimensions in pixels
cell_width = sheet.column_dimensions[column_letter].width * 7  # approx conversion factor
cell_height = sheet.row_dimensions[row_number].height

# Open the image using PIL (Pillow) to get its original dimensions
pil_img = PILImage.open(img_path)
original_width, original_height = pil_img.size

# Calculate the scaling factor
scale_x = cell_width / original_width
scale_y = cell_height / original_height

# Apply the minimum scale to keep the aspect ratio intact
scale = min(scale_x, scale_y)

# Resize the image using the scale
img.width = int(original_width * scale)
img.height = int(original_height * scale)

# Add the image to a specific cell
sheet.add_image(img, f'{column_letter}{row_number}')

# Save the workbook
workbook.save('images_scaled.xlsx')